# -*- coding:utf-8 _*-
#Author:xuehua
#Email:564033627@qq.com
#Date:2019/9/5
# Function：操作Excel练习

import openpyxl
from common import contants, http_request


class Case:
    """
    测试用例类，每个测试用例，实际就是一个实例
    """

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.exspected = None
        self.check_sql = None


class DoExcel:
    """
    操作Excel、获取测试数据，向Excel写入测试结果
    """

    def __init__(self, file_name, sheet_name):
        """
        初始化：打开表格，定位到表单
        :param file_name: 文件路径，文件名
        :param sheet_name: 表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        """
        获取表格中的测试用例
        :return:
        """
        max_row = self.sheet.max_row  # 获取最大行数
        cases = []  # 存储全部用例
        for r in range(2, max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(r, 1).value
            case.title = self.sheet.cell(r, 2).value
            case.url = self.sheet.cell(r, 3).value
            case.data = self.sheet.cell(r, 4).value
            case.method = self.sheet.cell(r, 5).value
            case.exspected = self.sheet.cell(r, 6).value
            case.check_sql = self.sheet.cell(r, 9).value
            cases.append(case)
        # 注意：读取完Excel表格后，记得关闭
        self.workbook.close()
        return cases

    def write_result(self, raw, actual, result):
        """
        将测试结果写入到Excel表格中
        :param raw:
        :param actual:
        :param result:
        :return:
        """
        sheet = self.workbook[self.sheet_name]
        sheet.cell(raw, 7).value = actual
        sheet.cell(raw, 8).value = result
        # 写入后，记得保存并关闭
        self.workbook.save(self.file_name)
        self.workbook.close()


if __name__ == '__main__':
    do_excel = DoExcel(contants.case_file, contants.sheet_name)
    http_request = http_request.HttpRequest()
    cases = do_excel.get_cases()
    # 执行每条用例，也就是遍历每条用例
    for case in cases:
        resp = http_request.request(method=case.method, url=case.url, data=case.data)
        print(resp.text)
        # print(type(resp))
        resp_dict = resp.json()  # 将响应报文转为字典
        # print(resp_dict)
        # print(type(case.exspected))#类
        actual = resp.text
        print(type(actual))
        if case.exspected == actual:
            do_excel.write_result(case.case_id + 1, actual, "PASS")
        else:
            do_excel.write_result(case.case_id + 1, actual, "Fail")